
from django.shortcuts import render, HttpResponse
from .models import *
import json

from django.http import JsonResponse,request
from django.core.serializers import serialize
from django.utils import timezone
from datetime import timedelta
from django.http import HttpResponseBadRequest

from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET, require_http_methods
from datetime import date,datetime, timedelta
from django.shortcuts import get_object_or_404
from calendar import monthrange
from django.db.models import Q
from datetime import datetime, timedelta
from .forms import DateRangeForm
from django.core.exceptions import ObjectDoesNotExist
from django.core.exceptions import ValidationError
from django import forms
from django.db import models

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
import configparser
import os
from django.shortcuts import render
import sys
import pandas as pd

from django.db.models import F
from django.http import HttpResponseServerError
import traceback


# Add the directory containing probe_agile_data to the Python path
sys.path.append(os.path.dirname(os.path.abspath(__file__)))



# Function for rendering the dashboard
# @require_http_methods(["GET"])
def rbinewhome(request): 
    try:
        # Get the current date
        end_date = timezone.now().date()
        # Get the start date by subtracting 6 days from the end date
        start_date = end_date - timedelta(days=6)
        
        # Combine database queries , List of source names
        source_names = ['rbi_fema', 'rbi_ecb', 'rbi_odi', 'startupIndia', 'sebi_ed_cgm', 'sebi_settlementorder', 'sebi_ao', 'sebi_chairperson_members', 'mca_roc', 'mca_rd',
                        'irdai_life_insurers', 'irdai_general_insurers', 'irdai_health_insurers', 'irdai_reinsurers', 'irdai_reinsurer_branches',
                         'irdai_corporate_surveyors', 'irdai_partner_surveyors', 'irdai_third_party_administrators','irdai_web_aggregators',
                        'irdai_insurance_repositories', 'irdai_insurance_marketing_firms', 'irdai_corporate_agents', 'irdai_telemarketer', 'pfrda_aggregators',  'pfrda_cra', 'pfrda_custodian',
                        'pfrda_pension_funds', 'pfrda_pop', 'pfrda_pop_se_npstrust', 'pfrda_ra_individual',  'pfrda_ra_renewal', 'pfrda_trustee_bank','cci_anti_profiteering_orders','cci_section31_formIII','cci_section43A_44', 
                        'nsdl_cp_issuance', 'nsdl_ncd_issuance', 'nsdl_cp_outstanding', 'nsdl_ncd_outstanding',
                        'nsdl_matured_securities_report', 'nsdl_active_securities_report','nsdl_isin_details', 'gem_suspended_sellers_entities','private_sector_companies_list_revised','ngo_trust_non_government_revised','ngo_academic_institutions_government_revised','ngo_academic_institutions_private_revised','ngo_other_registered_entities_non_government_revised','ngo_registered_societies_non_government_revised']
        
        # Initialize dictionaries to store data
        data = {}
        latest_counts = {}
        recent_statuses = {}
        recent_colors = {}
        
        # Loop through each source name
        for source_name in source_names:
            # Check which database the source belongs to
            if rbi_log.objects.using('rbi').filter(source_name=source_name).exists():
                # Use 'rbi' database
                db_name = 'rbi'
                model = rbi_log
            elif startup_india_log.objects.using('startup_india').filter(source_name=source_name).exists():
                db_name = 'startup_india'
                model = startup_india_log    
            elif sebi_log.objects.using('sebi').filter(source_name=source_name).exists():
                # Use 'sebi' database
                db_name = 'sebi'
                model = sebi_log
            elif mca_log.objects.using('mca').filter(source_name=source_name).exists():
                # Use 'sebi' database
                db_name = 'mca'
                model = mca_log
            elif irdai_log.objects.using('irdai').filter(source_name=source_name).exists():
                # Use 'sebi' database
                db_name = 'irdai'
                model = irdai_log
            elif pfrda_log.objects.using('pfrda').filter(source_name=source_name).exists():
                # Use 'sebi' database
                db_name = 'pfrda'
                model = pfrda_log
            elif cci_log.objects.using('cci').filter(source_name=source_name).exists():
                # Use 'cci' database
                db_name = 'cci'
                model = cci_log  
            elif nsdl_log.objects.using('nsdl').filter(source_name=source_name).exists():
                # Use 'cci' database
                db_name = 'nsdl'
                model = nsdl_log  
            elif gem_log.objects.using('gem').filter(source_name=source_name).exists():
                # Use 'cci' database
                db_name = 'gem'
                model = gem_log 
            elif ngo_log.objects.using('ngo').filter(source_name=source_name).exists():
                # Use 'cci' database
                db_name = 'ngo'
                model = ngo_log         

            else:
                # Handle other cases or raise an error
                continue  # or raise an error
            
            # Query the database for data within the specified date range
            data[source_name] = model.objects.using(db_name).filter(source_name=source_name, date_of_scraping__date__range=[start_date, end_date]).order_by('-date_of_scraping')
            
            # Get the latest entry for the source status
            latest_entry = get_latest_entry(db_name, source_name, model)
            
            # Get the latest total count for the source
            latest_counts[source_name] = get_latest_count(latest_entry)
            
            # Get the recent status for the source status
            recent_statuses[source_name] = latest_entry.source_status if latest_entry else 'N/A'
            
            # Get the color for the recent status of the source
            recent_colors[source_name] = get_color_for_source_status(recent_statuses[source_name])
        
        # Prepare data list for rendering in the template
        data_list = []
        for date in (end_date - timedelta(days=i) for i in range(7)):
            data_entry = {}
            for source_name in source_names:
                if source_name not in data:
                    continue
                entry = data[source_name].filter(date_of_scraping__date=date).first()
                data_available, data_scraped, status, reason = get_entry_info(entry)
                color = get_color(status, reason)
                data_entry[f'{source_name}_Data_Available'] = data_available
                data_entry[f'{source_name}_Data_Scraped'] = data_scraped
                data_entry[f'{source_name}_Color'] = color
            data_entry['Date'] = date.strftime('%d-%m-%Y')
            data_list.append(data_entry)
        # print("data_list===", data_list) 
        
        # Prepare context for rendering the template
        context = {
            'data_list': data_list, 
            'latest_counts': latest_counts, 
            'recent_statuses': recent_statuses,
            'recent_colors': recent_colors,
        }

        
       

        # Convert data to DataFrame for printing value in terminal
        data_df = pd.DataFrame(data_list)
        counts_df = pd.DataFrame(latest_counts.items(), columns=['Source', 'Latest Count'])
        statuses_df = pd.DataFrame(recent_statuses.items(), columns=['Source', 'Recent Status'])
        colors_df = pd.DataFrame(recent_colors.items(), columns=['Source', 'Recent Color'])

    
        # Set display options to show all columns in terminal
        pd.set_option('display.max_columns', None)

        # Print the DataFrames in terminal
        print("Data List DataFrame:")
        
        # print(data_df.to_string(index=False, header=True))
        # print("\nLatest Counts DataFrame:")
        # print(counts_df.to_string(index=False, header=True))
        # print("\nRecent Statuses DataFrame:")
        # print(statuses_df.to_string(index=False, header=True))
        # print("\nRecent Colors DataFrame:")
        # print(colors_df.to_string(index=False, header=True))
        
        # Convert DataFrame to HTML to see html page 
        data_html = data_df.to_html()
        counts_html = counts_df.to_html()
        statuses_html = statuses_df.to_html()
        colors_html = colors_df.to_html()

        # Print HTML tables in terminal 
        # print("Data List HTML Table:")
        # print(data_html)
        # print("\nLatest Counts HTML Table:")
        # print(counts_html)
        # print("\nRecent Statuses HTML Table:")
        # print(statuses_html)
        # print("\nRecent Colors HTML Table:")
        # print(colors_html)
        
        
        return render(request, 'fema/grid_dashboard.html', context)
    
    
        # # Convert the data to DataFrames
        # data_df = pd.DataFrame(data_list)
        # counts_df = pd.DataFrame(latest_counts.items(), columns=['Source', 'Latest Count'])
        # statuses_df = pd.DataFrame(recent_statuses.items(), columns=['Source', 'Recent Status'])
        # colors_df = pd.DataFrame(recent_colors.items(), columns=['Source', 'Recent Color'])

        # # Convert DataFrames to HTML tables
        # data_table = data_df.to_html(index=False, escape=False)
        # counts_table = counts_df.to_html(index=False, escape=False)
        # statuses_table = statuses_df.to_html(index=False, escape=False)
        # colors_table = colors_df.to_html(index=False, escape=False)

        # # Prepare the final response
        # response_data = {
        #     'data_table': data_table,
        #     'counts_table': counts_table,
        #     'statuses_table': statuses_table,
        #     'colors_table': colors_table
        # }

        # # Return the response as JSON
        # return JsonResponse(response_data)
    
    except Exception as e:
        # Log and print the exception
        traceback.print_exc()
        print("An exception occurred:", e)
        
        # Handle the exception and provide an appropriate response
        return HttpResponseServerError("An error occurred while rendering the dashboard. Please try again later.")


# Function to extract relevant information from a database entry and format it appropriately
def get_entry_info(entry):
    
    """
    Function to extract information from a database entry and format it appropriately.

    Parameters:
        entry: Database entry from which information needs to be extracted.

    Returns:
        Tuple: A tuple containing information extracted from the entry. The tuple structure is as follows:
            - data_available: The availability of data.
            - data_scraped: The amount of data scraped.
            - script_status: The status of the script.
            - failure_reason: The reason for failure, if any.
    """
    try:
        if entry:
            script_status = entry.script_status
            if script_status == 'not run':
                data_available = '0' if entry.data_available is None else entry.data_available
                data_scraped = '0' if entry.data_scraped is None else entry.data_scraped
            else:
                data_available = entry.data_available if entry.data_available is not None else '0' if entry.script_status == 'Success' else 'NA'
                data_scraped = entry.data_scraped if entry.data_scraped is not None else '0' if entry.script_status == 'Success' else 'NA'
            return data_available, data_scraped, script_status, entry.failure_reason
        return '-', '-', 'N/A', None
    except Exception as e:
        # Log and print the exception
        traceback.print_exc()
        print("Exception in get_entry_info:", e)
    

 # Get the latest entry for the source status    
def get_latest_entry(db_name, source_name, model):
    try:
        # Fetch the most recent entry for today's date
        recent_status_entry = model.objects.using(db_name).filter(source_name=source_name).latest('date_of_scraping')

        # If the status is not null, return the entry
        if recent_status_entry.source_status:
            return recent_status_entry

        # If today's status is null, retrieve all entries for today's date
        entries_for_today = model.objects.using(db_name).filter(source_name=source_name, date_of_scraping__date=timezone.now().date())

        # Check if there are multiple entries for today's date
        if entries_for_today.count() > 1:
            # Sort the entries based on date_of_scraping in descending order
            sorted_entries = entries_for_today.order_by('-date_of_scraping')
            for entry in sorted_entries:
                # Return the first entry with a non-null source_status
                if entry.source_status:
                    return entry

        # If today's status is still null or there is only one entry, iterate over previous dates until a non-null status is found
        for i in range(1, 100):  # Assuming checking for the past 99 days
            previous_date = timezone.now().date() - timedelta(days=i)
            previous_status_entry = model.objects.using(db_name).filter(source_name=source_name, date_of_scraping__date=previous_date).first()
            if previous_status_entry and previous_status_entry.source_status:
                return previous_status_entry

    except ObjectDoesNotExist:
        pass
    except Exception as e:
        # Log and print the exception
        traceback.print_exc()
        print("Exception in get_latest_entry:", e)
    return None  # Return None if no status is found


# Get the latest total count for the source
def get_latest_count(entry):
    try:
        if entry:
            if entry.total_record_count is not None:
                return str(entry.total_record_count)
            elif entry.script_status == 'Success':
                return '0'
        return '-'
    except Exception as e:
        # Log and print the exception
        traceback.print_exc()
        print("Exception in get_latest_count:", e)
   

#set the colour for data_available and data_scraped with respect to script status such as Success and failure 
def get_color(status, reason=None):
    color_map = {
        'Success': 'green',
        'Failure': 'red',
        'N/A': 'black'
    }
    if status == 'Failure' and reason and '204' in str(reason):
        return 'orange'
    return color_map.get(status, 'black')

# Get the color for the recent status of the source such as Active , Hibernated, Inactive
def get_color_for_source_status(status):
    if status == 'Active':
        return 'green'
    elif status == 'Hibernated':
        return 'orange'
    elif status == 'Inactive':
        return 'red'
    else:
        return 'black'



# @require_http_methods(["GET"])
def rbiget_data_for_popup1(request, source_name):
    try:
        today_date = timezone.now().date()
        
        # Initialize db_name and model variables
        db_name = None
        model = None
        
        # Check if the source name exists in 'rbi' database
        if rbi_log.objects.using('rbi').filter(source_name=source_name).exists():
            db_name = 'rbi'
            model = rbi_log

        elif startup_india_log.objects.using('startup_india').filter(source_name=source_name).exists():
            db_name = 'startup_india'
            model = startup_india_log    

            
        # Check if the source name exists in 'sebi' database
        elif sebi_log.objects.using('sebi').filter(source_name=source_name).exists():
            db_name = 'sebi'
            model = sebi_log
            
        # Check if the source name exists in 'mca_orders' database
        elif mca_log.objects.using('mca').filter(source_name=source_name).exists():
            db_name = 'mca'
            model = mca_log
            
        # Check if the source name exists in 'irdai' database
        elif irdai_log.objects.using('irdai').filter(source_name=source_name).exists():
            db_name = 'irdai'
            model = irdai_log
            
             # Check if the source name exists in 'pfrda' database
        elif pfrda_log.objects.using('pfrda').filter(source_name=source_name).exists():
            db_name = 'pfrda'
            model = pfrda_log

        # Check if the source name exists in 'cci' database
        elif cci_log.objects.using('cci').filter(source_name=source_name).exists():
            db_name = 'cci'
            model = cci_log  

        # Check if the source name exists in 'nsdl' database
        elif nsdl_log.objects.using('nsdl').filter(source_name=source_name).exists():
            db_name = 'nsdl'
            model = nsdl_log  

        # Check if the source name exists in 'gem' database
        elif gem_log.objects.using('gem').filter(source_name=source_name).exists():
            db_name = 'gem'
            model = gem_log  

        elif ngo_log.objects.using('ngo').filter(source_name=source_name).exists():
            db_name = 'ngo'
            model = ngo_log                
        else:
            # Handle other cases or raise an error
            return HttpResponse(status=404)

        
        # Query the appropriate database and model
        data = model.objects.using(db_name)\
            .filter(source_name=source_name, date_of_scraping__date=today_date)\
            .order_by('-date_of_scraping')\
            .first()

        if data:
            # Replace None values with hyphen
            data_scraped = data.data_scraped if data.data_scraped is not None else "0"
            failure_reason = data.failure_reason if data.failure_reason is not None else "-"

            if hasattr(data, 'deleted_source_count'):
                deleted_source_count = data.deleted_source_count if data.deleted_source_count else "0"
            else:
                deleted_source_count = "-"

            # Handle 'newly_added_count' to be '0' if it's None or doesn't exist
            # Handle 'newly_added_count' with three conditions
            if hasattr(data, 'newly_added_count'):
                newly_added_count = data.newly_added_count if data.newly_added_count else "0"
            else:
                newly_added_count = "-"


            response_data = {
                'source_name': data.source_name,
                'script_status': data.script_status,
                'data_scraped': data_scraped,
                'failure_reason': failure_reason,
                'date_of_scraping': data.date_of_scraping.strftime('%d-%m-%Y'),
                'deleted_source_count': deleted_source_count,
                'newly_added_count': newly_added_count,
            }
            print("response_data====", response_data)
            
            json_results= json.dumps(response_data) 
            
            return HttpResponse(json_results, content_type="application/json")
            
        else:
            return HttpResponse(status=404)
          
    except Exception as e:
        # Log and print the exception
        traceback.print_exc()
        print("Exception in rbiget_data_for_popup1:", e)
        return HttpResponseServerError("An error occurred while retrieving data. Please try again later.")


###################################################################################################################################################




def get_config_path(source_name):
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    config_folder = os.path.join(base_dir, 'config')
    config_file = f'config_{source_name.lower()}.ini'
    return os.path.join(config_folder, config_file)


def read_config(source_name):
    config_path = get_config_path(source_name)
    
    config = configparser.ConfigParser()
    config.read(config_path)

    return config.get(source_name, 'status')




"""Format a date to a string in 'dd-mm-YYYY' format."""
def format_date(date):
    return date.strftime('%d-%m-%Y') if date else ''

"""Get default start and end dates for a date range of past 7 days."""
def get_default_start_end_dates():
    end_date = datetime.now().date()
    start_date = end_date - timedelta(days=6)  # Default to past 7 days
    return start_date, end_date

"""Get start and end dates for a date range of past 15 days."""
def get_past_15_days():
    end_date = datetime.now().date()
    start_date = end_date - timedelta(days=14)
    return start_date, end_date

"""Get start and end dates for a date range of past month (30 days)."""
def get_past_month():
    today = datetime.now().date()
    end_date = today
    start_date = today - timedelta(days=29)  # Modify to get data for the past 30 days, including today
    return start_date, end_date

#  This code is for past month results .for example current month is december past month results is november month.
# def get_last_month():
#     today = datetime.now().date()
#     end_date = today.replace(day=1) - timedelta(days=1)
#     start_date = end_date.replace(day=1)
#     return start_date, end_date

"""Get color based on script_status and failure_reason."""
def get_status_color(script_status, failure_reason):
    if script_status == 'Success':
        return 'green'
    elif script_status == 'Failure' and '204' in str(failure_reason):
        return 'orange'
    else:
        return 'red'


"""navigation page , Filter and process data based on date range and source name. dropdown functionality included """
# @require_http_methods(["GET"])
def filter_data(request, source_name, model, db_name):
    """Filter and process data based on date range and source name."""
    try:
        form = DateRangeForm(request.GET)
        
        # Default values for start_date and end_date
        start_date, end_date = get_default_start_end_dates()
        
        if form.is_valid():
            date_range = form.cleaned_data.get('date_range')
            if date_range == 'past_7_days':
                start_date, end_date = get_default_start_end_dates()
            elif date_range == 'past_15_days':
                start_date, end_date = get_past_15_days()
            elif date_range == 'past_month':
                start_date, end_date = get_past_month()
            elif date_range == 'custom':
                start_date = form.cleaned_data.get('start_date')
                end_date = form.cleaned_data.get('end_date')
                if start_date and end_date:
                    date_difference = end_date - start_date
                    if date_difference.days > 60:
                        # Adjust end_date if it's more than 60 days from start_date
                        end_date = start_date + timedelta(days=60)
            else:
                start_date, end_date = get_default_start_end_dates()

        # Adjust end_date to cover the entire day
        # end_date = end_date + timedelta(days=1)
        
        
        # Query the database with the adjusted date range
        unique_dates = model.objects.using(db_name).filter(
            date_of_scraping__date__range=[start_date, end_date],
            source_name=source_name
            ).values('date_of_scraping__date').distinct().order_by('-date_of_scraping__date')

        formatted_data = []
        for unique_date in unique_dates:
            latest_entry = model.objects.using(db_name).filter(
                source_name=source_name,
                date_of_scraping__date=unique_date['date_of_scraping__date']
            ).order_by('-date_of_scraping').first()  # Use first() instead of latest() for sorting in descending order

            formatted_date = format_date(latest_entry.date_of_scraping)
            status_color = get_status_color(latest_entry.script_status, latest_entry.failure_reason)

            data_available = latest_entry.data_available if latest_entry.data_available is not None else "0"
            data_scraped = latest_entry.data_scraped if latest_entry.data_scraped is not None else "0"
            failure_reason = latest_entry.failure_reason if latest_entry.failure_reason is not None else "-"

            
            if hasattr(latest_entry, 'deleted_source_count'):
                deleted_source_count = latest_entry.deleted_source_count if latest_entry.deleted_source_count else "0"
            else:
                deleted_source_count = "-"

            # Handle 'newly_added_count' to be '0' if it's None or doesn't exist
            # Handle 'newly_added_count' with three conditions
            if hasattr(latest_entry, 'newly_added_count'):
                newly_added_count = latest_entry.newly_added_count if latest_entry.newly_added_count else "0"
            else:
                newly_added_count = "-"


            formatted_data.append({
                'source_name': latest_entry.source_name,
                'script_status': latest_entry.script_status,
                'failure_reason': failure_reason,
                'data_available': data_available,
                'data_scraped': data_scraped,
                'date_of_scraping': formatted_date,
                'status_color': status_color,
                'deleted_source_count': deleted_source_count,
                'newly_added_count': newly_added_count,
            })
            
        
        # Handle export to Excel functionality
        if 'download' in request.GET:
            date_range = request.GET.get('date_range', 'past_7_days')
            start_date = request.GET.get('start_date')
            end_date = request.GET.get('end_date')

            # Call the export_to_excel function with the selected date range
            return export_to_excel(request, formatted_data, date_range, start_date, end_date, source_name, model, db_name)
        
        
        recent_status, recent_color = get_recent_status_and_color(source_name, model, db_name)

        
        # Prepare context for rendering the template
        context = {
            'form': form,
            'data': formatted_data,
            'start_date': format_date(start_date),
            'end_date': format_date(end_date),
            'past_15_days': (format_date(get_past_15_days()[0]), format_date(get_past_15_days()[1])),
            'last_month': (format_date(get_past_month()[0]), format_date(get_past_month()[1])),
            'source_name': source_name,
            'recent_status': recent_status,
            'recent_color': recent_color,
        
        }

        return render(request, 'fema/grid_datefilter.html', context)
    
    except Exception as e:
        # Log and print the exception
        traceback.print_exc()
        print("Exception in filter_data:", e)
        return HttpResponseServerError("An error occurred while processing the data.")




# rbi urls

def rbinewfema_datefilter(request):
    return filter_data(request, 'rbi_fema', rbi_log, 'rbi')

def rbinewecb_datefilter(request):
    return filter_data(request, 'rbi_ecb', rbi_log, 'rbi')

def rbinewodi_datefilter(request):
    return filter_data(request, 'rbi_odi',  rbi_log, 'rbi')

def startupindia_datefilter(request):
    return filter_data(request, 'startupIndia', startup_india_log, 'startup_india')

# sebi urls

def sebi_ed_datefilter(request):
    return filter_data(request, 'sebi_ed_cgm', sebi_log, 'sebi')

def sebi_so_datefilter(request):
    return filter_data(request, 'sebi_settlementorder', sebi_log, 'sebi')

def sebi_ao_datefilter(request):
    return filter_data(request, 'sebi_ao',  sebi_log, 'sebi')

def sebi_members_datefilter(request):
    return filter_data(request, 'sebi_chairperson_members', sebi_log, 'sebi')


# mca urls

def mca_roc_datefilter(request):
    return filter_data(request, 'mca_roc', mca_log, 'mca')


def mca_rd_datefilter(request):
    return filter_data(request, 'mca_rd', mca_log, 'mca')

# irdai urls


def irdai_life_insurers_datefilter(request):
    return filter_data(request, 'irdai_life_insurers', irdai_log, 'irdai')

def irdai_general_insurers_datefilter(request):
    return filter_data(request, 'irdai_general_insurers', irdai_log, 'irdai')

def irdai_health_insurers_datefilter(request):
    return filter_data(request, 'irdai_health_insurers', irdai_log, 'irdai')

def irdai_reinsurers_datefilter(request):
    return filter_data(request, 'irdai_reinsurers', irdai_log, 'irdai')

def irdai_reinsurer_branches_datefilter(request):
    return filter_data(request, 'irdai_reinsurer_branches', irdai_log, 'irdai')

def irdai_corporate_surveyors_datefilter(request):
    return filter_data(request, 'irdai_corporate_surveyors', irdai_log, 'irdai')

def irdai_partner_surveyors_datefilter(request):
    return filter_data(request, 'irdai_partner_surveyors', irdai_log, 'irdai')

def irdai_third_party_administrators_datefilter(request):
    return filter_data(request, 'irdai_third_party_administrators', irdai_log, 'irdai')

def irdai_web_aggregators_datefilter(request):
    return filter_data(request, 'irdai_web_aggregators', irdai_log, 'irdai')

def irdai_insurance_repositories_datefilter(request):
    return filter_data(request, 'irdai_insurance_repositories', irdai_log, 'irdai')

def irdai_insurance_marketing_firms_datefilter(request):
    return filter_data(request, 'irdai_insurance_marketing_firms', irdai_log, 'irdai')

def irdai_corporate_agents_datefilter(request):
     return filter_data(request, 'irdai_corporate_agents', irdai_log, 'irdai')

def irdai_telemarketer_datefilter(request):
    return filter_data(request, 'irdai_telemarketer', irdai_log, 'irdai')

# pfrda urls

def pfrda_aggregators_datefilter(request):
    return filter_data(request, 'pfrda_aggregators', pfrda_log, 'pfrda')

def pfrda_cra_datefilter(request):
    return filter_data(request, 'pfrda_cra', pfrda_log, 'pfrda')

def pfrda_custodian_datefilter(request):
    return filter_data(request, 'pfrda_custodian', pfrda_log, 'pfrda')

def pfrda_pension_funds_datefilter(request):
    return filter_data(request, 'pfrda_pension_funds', pfrda_log, 'pfrda')

def pfrda_pop_datefilter(request):
    return filter_data(request, 'pfrda_pop', pfrda_log, 'pfrda')

def pfrda_pop_se_datefilter(request):
    return filter_data(request, 'pfrda_pop_se_npstrust', pfrda_log, 'pfrda')

def pfrda_ra_individual_datefilter(request):
    return filter_data(request, 'pfrda_ra_individual', pfrda_log, 'pfrda')

def pfrda_ra_renewal_datefilter(request):
    return filter_data(request, 'pfrda_ra_renewal', pfrda_log, 'pfrda')

def pfrda_trustee_bank_datefilter(request):
    return filter_data(request, 'pfrda_trustee_bank', pfrda_log, 'pfrda')

def cci_anti_profiteering_orders_datefilter(request):
    return filter_data(request, 'cci_anti_profiteering_orders', cci_log, 'cci')

def cci_section31_formIII_datefilter(request):
    return filter_data(request, 'cci_section31_formIII', cci_log, 'cci')

def cci_section43A_44_datefilter(request):
    return filter_data(request, 'cci_section43a_44', cci_log, 'cci')

# NSDL URLS
def nsdl_cp_issuance_datefilter(request):
    return filter_data(request, 'nsdl_cp_issuance', nsdl_log, 'nsdl')

def nsdl_ncd_issuance_datefilter(request):
    return filter_data(request, 'nsdl_ncd_issuance', nsdl_log, 'nsdl')

def nsdl_cp_outstanding_datefilter(request):
    return filter_data(request, 'nsdl_cp_outstanding', nsdl_log, 'nsdl')

def nsdl_ncd_outstanding_datefilter(request):
    return filter_data(request, 'nsdl_ncd_outstanding', nsdl_log, 'nsdl')

def nsdl_matured_securities_report_datefilter(request):
    return filter_data(request, 'nsdl_matured_securities_report', nsdl_log, 'nsdl')

def nsdl_active_securities_report_datefilter(request):
    return filter_data(request, 'nsdl_active_securities_report', nsdl_log, 'nsdl')

def nsdl_isin_details_datefilter(request):
    return filter_data(request, 'nsdl_isin_details', nsdl_log, 'nsdl')

# GEM URLS
def gem_suspended_sellers_entities_datefilter(request):
    return filter_data(request, 'gem_suspended_sellers_entities', gem_log, 'gem')

def ngo_private_sector_companies_datefilter(request):
    return filter_data(request, 'private_sector_companies_list_revised', ngo_log, 'ngo')
def ngo_trust_non_government_datefilter(request):
    return filter_data(request, 'ngo_trust_non_government_revised', ngo_log, 'ngo')
def ngo_academic_institutions_government_datefilter(request):
    return filter_data(request, 'ngo_academic_institutions_government_revised', ngo_log, 'ngo')
def ngo_academic_institutions_private_datefilter(request):
    return filter_data(request, 'ngo_academic_institutions_private_revised', ngo_log, 'ngo')
def ngo_other_registered_entities_non_government_datefilter(request):
    return filter_data(request, 'ngo_other_registered_entities_non_government_revised', ngo_log, 'ngo')
def ngo_registered_societies_non_government_datefilter(request):
    return filter_data(request, 'ngo_registered_societies_non_government_revised', ngo_log, 'ngo')

# startup india




"""Get the recent status and color for a given source_name in navigation page."""
def get_recent_status_and_color(source_name, model, db_name):
    try:
        recent_status_entry = model.objects.using(db_name).filter(source_name=source_name).latest('date_of_scraping')
        recent_status = recent_status_entry.source_status if recent_status_entry is not None and recent_status_entry.source_status else None
        if recent_status is not None:
            status_color = (
                'greenyellow' if recent_status == 'Active' else
                'orange' if recent_status == 'Hibernated' else
                'red' if recent_status == 'Inactive' else
                'black'
            )
            return recent_status, status_color

        for i in range(1, 7):  # Assuming checking for the past 6 days
            previous_date = timezone.now().date() - timedelta(days=i)
            previous_status_entry = model.objects.using(db_name).filter(source_name=source_name, date_of_scraping__date=previous_date).first()
            if previous_status_entry is not None and previous_status_entry.source_status:
                status_color = get_status_color(previous_status_entry.script_status, previous_status_entry.failure_reason)
                return previous_status_entry.source_status, status_color

    except ObjectDoesNotExist:
        pass
    except Exception as e:
        # Log and print the exception
        traceback.print_exc()
        print("Exception in get_recent_status_and_color:", e)
    return 'N/A', 'black'


""" funcionality for Export data to an Excel file."""
def export_to_excel(request, data, date_range, start_date, end_date, source_name, model, db_name):
    """Export data to an Excel file."""
    try:
        if date_range == 'past_15_days':
            end_date = datetime.now().date()
            start_date = end_date - timedelta(days=14)
        elif date_range == 'past_month':
            today = datetime.now().date()
            end_date = today
            start_date = today - timedelta(days=29)
        elif date_range == 'custom':
            # Add logic to handle custom view start_date and end_date
            if start_date and end_date:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            else:
                # Default to past 7 days if no specific range is selected
                end_date = datetime.now().date()
                start_date = end_date - timedelta(days=6)
        else:
            # Default to past 7 days if no specific range is selected
            end_date = datetime.now().date()
            start_date = end_date - timedelta(days=6)
            
        # Get the latest data for each unique date
        unique_dates = model.objects.using(db_name).filter(
            date_of_scraping__date__range=[start_date, end_date],
            source_name=source_name
        ).values('date_of_scraping__date').distinct()

        latest_data = []
        for unique_date in unique_dates:
            latest_entry = model.objects.using(db_name).filter(
                source_name=source_name,
                date_of_scraping__date=unique_date['date_of_scraping__date']
            ).order_by('-date_of_scraping').first()

            if latest_entry:
                latest_data.append(latest_entry)

        # Generate a dynamic filename based on source name and date range
        filename = f"{source_name}_{start_date}_{end_date}.xlsx"

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={filename}'

        # Generate Excel file
        workbook = Workbook()
        worksheet = workbook.active

        # Add headers to the worksheet
        header_font = Font(bold=True)
        headers = ['Source Name', 'Status', '#Records Available', '#Records Scraped', 'Failure Reason', 'Scraped On']

        for col_num, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            
        # Populate data in the worksheet
        latest_data.reverse()  # Reverse the order to go from latest to previous

        # Populate data in the worksheet
        for row_num, data_entry in enumerate(latest_data, start=2):
            worksheet.cell(row=row_num, column=1, value=data_entry.source_name)
            worksheet.cell(row=row_num, column=2, value=data_entry.script_status)
            worksheet.cell(row=row_num, column=3, value=data_entry.data_available if data_entry.data_available is not None else "0")
            worksheet.cell(row=row_num, column=4, value=data_entry.data_scraped if data_entry.data_scraped is not None else "0")
            worksheet.cell(row=row_num, column=5, value=data_entry.failure_reason or "-")
            worksheet.cell(row=row_num, column=6, value=format_date(data_entry.date_of_scraping))

        # Save the workbook and prepare the response for download
        workbook.save(response)
        return response
    except Exception as e:
        # Log and print the exception
        traceback.print_exc()
        print("Exception in export_to_excel:", e)
        return HttpResponseServerError("An error occurred while exporting data to Excel.")
    

###################################################################################################################################################



def rbi_tab(request):
    startup_data= startup_india_log.objects.using('startup_india').all()
    return render(request,'fema/index.html', {'startup_data':startup_data}) 


# #function for dashboard functionality
# def rbinewhome(request): 
    
#     # Get the current date and the date 6 days ago
#     end_date = timezone.now().date()
#     start_date = end_date - timedelta(days=6)
    
#     # Query the database for rbi_fema , rbi_ecb ,rbi_odi, startupindia data within the last 7 days
#     fema_data = rbi_log.objects.using('rbi').filter(source_name='rbi_fema', date_of_scraping__date__range=[start_date, end_date]).order_by('-date_of_scraping')
#     ecb_data = rbi_log.objects.using('rbi').filter(source_name='rbi_ecb', date_of_scraping__date__range=[start_date, end_date]).order_by('-date_of_scraping')
#     odi_data = rbi_log.objects.using('rbi').filter(source_name='rbi_odi', date_of_scraping__date__range=[start_date, end_date]).order_by('-date_of_scraping')
    
#     startupindia_data = rbi_log.objects.using('rbi').filter(source_name='startupindia', date_of_scraping__date__range=[start_date, end_date]).order_by('-date_of_scraping')
   
    
#     # Try to get the latest total_record_count for rbi_fema and rbi_ecb, rbi_odi, startupindia
#     try:
#         fema_latest_entry = rbi_log.objects.using('rbi').filter(source_name='rbi_fema').latest('date_of_scraping')
#         # Set fema_latest_count to 0 for success, '-' for failure, or '-' if total_record_count is None
#         fema_latest_count = fema_latest_entry.total_record_count if fema_latest_entry.total_record_count is not None else "0" if fema_latest_entry.script_status == 'Success' else "-"
#     except ObjectDoesNotExist:
#         # Set fema_latest_count to '-' if there is no latest entry
#         fema_latest_count = "-"

#     try:
#         ecb_latest_entry = rbi_log.objects.using('rbi').filter(source_name='rbi_ecb').latest('date_of_scraping')
#         # Set ecb_latest_count to 0 for success, '-' for failure, or '-' if total_record_count is None
#         ecb_latest_count = ecb_latest_entry.total_record_count if ecb_latest_entry.total_record_count is not None else "0" if ecb_latest_entry.script_status == 'Success' else "-"
#     except ObjectDoesNotExist:
#         # Set ecb_latest_count to '-' if there is no latest entry
#         ecb_latest_count = "-"
        
#     try:
#         odi_latest_entry = rbi_log.objects.using('rbi').filter(source_name='rbi_odi').latest('date_of_scraping')
#         odi_latest_count = odi_latest_entry.total_record_count if odi_latest_entry.total_record_count is not None else "0" if odi_latest_entry.script_status == 'Success' else "-"
#     except ObjectDoesNotExist:
#         odi_latest_count = "-"
   
#     try:
#         startupindia_latest_entry = rbi_log.objects.using('rbi').filter(source_name='startupindia').latest('date_of_scraping')
#         startupindia_latest_count = startupindia_latest_entry.total_record_count if startupindia_latest_entry.total_record_count is not None else "0" if startupindia_latest_entry.script_status == 'Success' else "-"
#     except ObjectDoesNotExist:
#         startupindia_latest_count = "-"

        
#     data_list = []
    
#     # Iterate over the last 7 days
#     for date in (end_date - timedelta(days=i) for i in range(7)):
#         # Get the rbi_fema entry and rbi_ecb entry ,rbi_odi,startupindia for the current date
#         fema_entry = fema_data.filter(date_of_scraping__date=date).first()
#         ecb_entry = ecb_data.filter(date_of_scraping__date=date).first()
#         odi_entry = odi_data.filter(date_of_scraping__date=date).first()
        
#         startupindia_entry = startupindia_data.filter(date_of_scraping__date=date).first()
        
#         # Set fema_data_available and fema_data_scraped to '0' for success, 'NA' for failure, or '-' if data is None
#         fema_data_available = fema_entry.data_available if fema_entry is not None and fema_entry.data_available is not None else "0" if fema_entry is not None and fema_entry.script_status == 'Success' else "NA" if fema_entry is not None and fema_entry.script_status == 'Failure' else "-"
#         fema_data_scraped = fema_entry.data_scraped if fema_entry is not None and fema_entry.data_scraped is not None else "0" if fema_entry is not None and fema_entry.script_status == 'Success' else "NA" if fema_entry is not None and fema_entry.script_status == 'Failure' else "-"
       
#         # Set ecb_data_available and ecb_data_scraped to '0' for success, 'NA' for failure, or '-' if data is None
#         ecb_data_available = ecb_entry.data_available if ecb_entry is not None and ecb_entry.data_available is not None else "0" if ecb_entry is not None and ecb_entry.script_status == 'Success' else "NA" if ecb_entry is not None and ecb_entry.script_status == 'Failure' else "-"
#         ecb_data_scraped = ecb_entry.data_scraped if ecb_entry is not None and ecb_entry.data_scraped is not None else "0" if ecb_entry is not None and ecb_entry.script_status == 'Success' else "NA" if ecb_entry is not None and ecb_entry.script_status == 'Failure' else "-"
       
        
#         odi_data_available = odi_entry.data_available if odi_entry is not None and odi_entry.data_available is not None else "0" if odi_entry is not None and odi_entry.script_status == 'Success' else "NA" if odi_entry is not None and odi_entry.script_status == 'Failure' else "-"
#         odi_data_scraped = odi_entry.data_scraped if odi_entry is not None and odi_entry.data_scraped is not None else "0" if odi_entry is not None and odi_entry.script_status == 'Success' else "NA" if odi_entry is not None and odi_entry.script_status == 'Failure' else "-"
        
#         startupindia_data_available = startupindia_entry.data_available if startupindia_entry is not None and startupindia_entry.data_available is not None else "0" if startupindia_entry is not None and startupindia_entry.script_status == 'Success' else "NA" if startupindia_entry is not None and startupindia_entry.script_status == 'Failure' else "-"
#         startupindia_data_scraped = startupindia_entry.data_scraped if startupindia_entry is not None and startupindia_entry.data_scraped is not None else "0" if startupindia_entry is not None and startupindia_entry.script_status == 'Success' else "NA" if startupindia_entry is not None and startupindia_entry.script_status == 'Failure' else "-"
       
         
#         fema_status = fema_entry.script_status if fema_entry is not None else 'N/A'
#         ecb_status = ecb_entry.script_status if ecb_entry is not None else 'N/A'
#         odi_status = odi_entry.script_status if odi_entry is not None else 'N/A'
#         startupindia_status = startupindia_entry.script_status if startupindia_entry is not None else 'N/A'
        
        
#         fema_reason = fema_entry.failure_reason if fema_entry is not None else None
#         ecb_reason = ecb_entry.failure_reason if ecb_entry is not None else None
#         odi_reason = odi_entry.failure_reason if odi_entry is not None else None
#         startupindia_reason = startupindia_entry.failure_reason if startupindia_entry is not None else None
        
        

#         # Determine the color based on status and reason
#         fema_color = ( 'green' if fema_status == 'Success' else 'orange' if fema_status == 'Failure' and '204' in str(fema_reason) else 'red' if fema_status == 'Failure' else 'black')
  
#         ecb_color = ( 'green' if ecb_status == 'Success' else 'orange' if ecb_status == 'Failure' and '204' in str(ecb_reason) else  'red' if ecb_status == 'Failure' else 'black')  
        
#         odi_color = ('green' if odi_status == 'Success' else 'orange' if odi_status == 'Failure' and '204' in str(odi_reason) else 'red' if odi_status == 'Failure' else 'black')
        
#         startupindia_color = ('green' if startupindia_status == 'Success' else 'orange' if startupindia_status == 'Failure' and '204' in str(startupindia_reason) else 'red' if startupindia_status == 'Failure' else 'black')
        
        
#         # Append data to the data_list for rendering in HTML
#         data_list.append({
#             'Date': date.strftime('%d-%m-%Y'),
#             'FEMA_Data_Available': fema_data_available,
#             'FEMA_Data_Scraped': fema_data_scraped,
#             'FEMA_Color': fema_color,
#             'ECB_Data_Available': ecb_data_available,
#             'ECB_Data_Scraped': ecb_data_scraped,
#             'ECB_Color': ecb_color,
#             'ODI_Data_Available': odi_data_available,
#             'ODI_Data_Scraped': odi_data_scraped,
#             'ODI_Color': odi_color,
#             'STARTUPINDIA_Data_Available' : startupindia_data_available,
#             'STARTUPINDIA_Data_Scraped' :startupindia_data_scraped,
#             'STARTUPINDIA_Color':startupindia_color,
#         })
    
    
#     # Fetch the recent status for each source_name
#     fema_recent_status = get_recent_status('rbi_fema')
#     ecb_recent_status = get_recent_status('rbi_ecb')
#     odi_recent_status = get_recent_status('rbi_odi')
#     startupindia_recent_status = get_recent_status('startupindia')

#     # Determine the color based on the recent source_status
#     fema_recent_color = ('green' if fema_recent_status == 'Active' else'orange' if fema_recent_status == 'Hibernated' else 'red' if fema_recent_status == 'Inactive' else 'black')

#     ecb_recent_color = ('green' if ecb_recent_status == 'Active' else 'orange' if ecb_recent_status == 'Hibernated' else 'red' if ecb_recent_status == 'Inactive' else 'black')

#     odi_recent_color = ('green' if odi_recent_status == 'Active' else 'orange' if odi_recent_status == 'Hibernated' else 'red' if odi_recent_status == 'Inactive' else 'black')

#     startupindia_recent_color = ('green' if startupindia_recent_status == 'Active' else 'orange' if startupindia_recent_status == 'Hibernated' else 'red' if startupindia_recent_status == 'Inactive' else 'black')

            
#     # # Update the source_status in the database
#     # update_status_in_database()
    
#     # Prepare the context to be passed to the HTML template
#     context = {
#         'data_list': data_list, 
#         'fema_latest_count': fema_latest_count, 
#         'ecb_latest_count': ecb_latest_count, 
#         'odi_latest_count': odi_latest_count, 
#         'startupindia_latest_count': startupindia_latest_count,
#         'fema_recent_status': fema_recent_status,
#         'ecb_recent_status': ecb_recent_status,
#         'odi_recent_status': odi_recent_status,
#         'startupindia_recent_status': startupindia_recent_status,
#         'fema_recent_color': fema_recent_color,
#         'ecb_recent_color': ecb_recent_color,
#         'odi_recent_color': odi_recent_color,
#         'startupindia_recent_color': startupindia_recent_color,
            
#     }
    
#     # Render the HTML template with the context
#     return render(request, 'fema/grid.html', context)







  
# def rbinewfema_datefilter(request):
#     form = DateRangeForm(request.GET)
    
#     # Default values for start_date and end_date
#     start_date, end_date = get_default_start_end_dates()
    
#     if form.is_valid():
#         date_range = form.cleaned_data.get('date_range')
#         if date_range == 'past_7_days':
#             start_date, end_date = get_default_start_end_dates()
#         elif date_range == 'past_15_days':
#             start_date, end_date = get_past_15_days()
#         elif date_range == 'last_month':
#             start_date, end_date = get_last_month()
#         elif date_range == 'custom':
#             start_date = form.cleaned_data.get('start_date')
#             end_date = form.cleaned_data.get('end_date')
#             if start_date and end_date:
#                 date_difference = end_date - start_date
#                 if date_difference.days > 60:
#                     # Adjust end_date if it's more than 60 days from start_date
#                     end_date = start_date + timedelta(days=60)
#         else:
#             start_date, end_date = get_default_start_end_dates()

#     data = rbi_log.objects.using('rbi').filter(
#         date_of_scraping__date__range=[start_date, end_date],
#         source_name='rbi_fema'
#     )

#     formatted_data = []
#     for item in data:
#         formatted_date = format_date(item.date_of_scraping)
#         status_color = get_status_color(item.script_status, item.failure_reason)
        
#         # Replace None values with hyphen
#         data_available = item.data_available if item.data_available is not None else "0"
#         data_scraped = item.data_scraped if item.data_scraped is not None else "0"
#         failure_reason = item.failure_reason if item.failure_reason is not None else "-"
        
#         formatted_data.append({
#             'source_name': item.source_name,
#             'script_status': item.script_status,
#             'failure_reason': failure_reason,
#             'data_available': data_available,
#             'data_scraped': data_scraped,
#             'date_of_scraping': formatted_date,
#             'status_color': status_color,
#         })

#     context = {
#         'form': form,
#         'data': formatted_data,
#         'start_date': format_date(start_date),
#         'end_date': format_date(end_date),
#         'past_15_days': (format_date(get_past_15_days()[0]), format_date(get_past_15_days()[1])),
#         'last_month': (format_date(get_last_month()[0]), format_date(get_last_month()[1])),
#         'table_name_filter': 'rbi_fema',
#         'source_name':'rbi_fema',
#     }

#     return render(request, 'fema/download_data.html', context)



# def rbinewecb_datefilter(request):
#     form = DateRangeForm(request.GET)
    
    
#     # Default values for start_date and end_date
#     start_date, end_date = get_default_start_end_dates()
    
#     if form.is_valid():
#         date_range = form.cleaned_data.get('date_range')
#         if date_range == 'past_7_days':
#             start_date, end_date = get_default_start_end_dates()
#         elif date_range == 'past_15_days':
#             start_date, end_date = get_past_15_days()
#         elif date_range == 'last_month':
#             start_date, end_date = get_last_month()
#         elif date_range == 'custom':
#             start_date = form.cleaned_data.get('start_date')
#             end_date = form.cleaned_data.get('end_date')
#             if start_date and end_date:
#                 date_difference = end_date - start_date
#                 if date_difference.days > 60:
#                     # Adjust end_date if it's more than 60 days from start_date
#                     end_date = start_date + timedelta(days=60)
#         else:
#             start_date, end_date = get_default_start_end_dates()

#     # Adjust end_date to cover the entire day
#     # end_date = end_date + timedelta(days=1)

#     data = rbi_log.objects.using('rbi').filter(
#          date_of_scraping__date__range=[start_date, end_date],
#         source_name='rbi_ecb'
#     )

#     formatted_data = []
#     for item in data:
#         formatted_date = format_date(item.date_of_scraping)
#         status_color = get_status_color(item.script_status, item.failure_reason)
        
#         # Replace None values with hyphen
#         data_available= item.data_available if item.data_available is not None else "0"
#         data_scraped = item.data_scraped if item.data_scraped is not None else "0"
#         failure_reason = item.failure_reason if item.failure_reason is not None else "-"
        
#         formatted_data.append({
#             'source_name': item.source_name,
#             'script_status': item.script_status,
#             'failure_reason':failure_reason,
#             'data_available':data_available,
#             'data_scraped': data_scraped,
#             'date_of_scraping': formatted_date,
#             'status_color': status_color,
#         })

#     context = {
#         'form': form,
#         'data': formatted_data,
#         'start_date': format_date(start_date),
#         'end_date': format_date(end_date),
#         'past_15_days': (format_date(get_past_15_days()[0]), format_date(get_past_15_days()[1])),
#         'last_month': (format_date(get_last_month()[0]), format_date(get_last_month()[1])),
#         'table_name_filter': 'rbi_ecb',
#         'source_name':'rbi_ecb',
#     }

#     return render(request, 'fema/download_data.html', context)



# def rbinewodi_datefilter(request):
#     form = DateRangeForm(request.GET)
    
    
#     # Default values for start_date and end_date
#     start_date, end_date = get_default_start_end_dates()
    
#     if form.is_valid():
#         date_range = form.cleaned_data.get('date_range')
#         if date_range == 'past_7_days':
#             start_date, end_date = get_default_start_end_dates()
#         elif date_range == 'past_15_days':
#             start_date, end_date = get_past_15_days()
#         elif date_range == 'last_month':
#             start_date, end_date = get_last_month()
#         elif date_range == 'custom':
#             start_date = form.cleaned_data.get('start_date')
#             end_date = form.cleaned_data.get('end_date')
#             if start_date and end_date:
#                 date_difference = end_date - start_date
#                 if date_difference.days > 60:
#                     # Adjust end_date if it's more than 60 days from start_date
#                     end_date = start_date + timedelta(days=60)
#         else:
#             start_date, end_date = get_default_start_end_dates()

#     # Adjust end_date to cover the entire day
#     # end_date = end_date + timedelta(days=1)

#     data = rbi_log.objects.using('rbi').filter(
#          date_of_scraping__date__range=[start_date, end_date],
#         source_name='rbi_odi'
#     )

#     formatted_data = []
#     for item in data:
#         formatted_date = format_date(item.date_of_scraping)
#         status_color = get_status_color(item.script_status, item.failure_reason)
        
#         # Replace None values with hyphen
#         data_available= item.data_available if item.data_available is not None else "0"
#         data_scraped = item.data_scraped if item.data_scraped is not None else "0"
#         failure_reason = item.failure_reason if item.failure_reason is not None else "-"
        
#         formatted_data.append({
#             'source_name': item.source_name,
#             'script_status': item.script_status,
#             'failure_reason':failure_reason,
#             'data_available':data_available,
#             'data_scraped': data_scraped,
#             'date_of_scraping': formatted_date,
#             'status_color': status_color,
#         })

#     context = {
#         'form': form,
#         'data': formatted_data,
#         'start_date': format_date(start_date),
#         'end_date': format_date(end_date),
#         'past_15_days': (format_date(get_past_15_days()[0]), format_date(get_past_15_days()[1])),
#         'last_month': (format_date(get_last_month()[0]), format_date(get_last_month()[1])),
#         'table_name_filter': 'rbi_odi',
#     }

#     return render(request, 'fema/gridfilter.html', context)


   
# def rbinewstartupindia_datefilter(request):
#     form = DateRangeForm(request.GET)
    
#      # Default values for start_date and end_date
#     start_date, end_date = get_default_start_end_dates()
    
#     if form.is_valid():
#         date_range = form.cleaned_data.get('date_range')
#         if date_range == 'past_7_days':
#             start_date, end_date = get_default_start_end_dates()
#         elif date_range == 'past_15_days':
#             start_date, end_date = get_past_15_days()
#         elif date_range == 'last_month':
#             start_date, end_date = get_last_month()
#         elif date_range == 'custom':
#             start_date = form.cleaned_data.get('start_date')
#             end_date = form.cleaned_data.get('end_date')
#             if start_date and end_date:
#                 date_difference = end_date - start_date
#                 if date_difference.days > 60:
#                     # Adjust end_date if it's more than 60 days from start_date
#                     end_date = start_date + timedelta(days=60)
#         else:
#             start_date, end_date = get_default_start_end_dates()

#     # Adjust end_date to cover the entire day
#     # end_date = end_date + timedelta(days=1)

#     data = rbi_log.objects.using('rbi').filter(
#          date_of_scraping__date__range=[start_date, end_date],
#         source_name='startupindia'
#     )

#     formatted_data = []
#     for item in data:
#         formatted_date = format_date(item.date_of_scraping)
#         status_color = get_status_color(item.script_status, item.failure_reason)
        
#          # Replace None values with hyphen
#         data_available= item.data_available if item.data_available is not None else "0"
#         data_scraped = item.data_scraped if item.data_scraped is not None else "0"
#         failure_reason = item.failure_reason if item.failure_reason is not None else "-"
        
#         formatted_data.append({
#             'source_name': item.source_name,
#             'script_status': item.script_status,
#             'failure_reason': failure_reason,
#             'data_available': data_available,
#             'data_scraped': data_scraped,
#             'date_of_scraping': formatted_date,
#             'status_color': status_color,
#         })

#     context = {
#         'form': form,
#         'data': formatted_data,
#         'start_date': format_date(start_date),
#         'end_date': format_date(end_date),
#         'past_15_days': (format_date(get_past_15_days()[0]), format_date(get_past_15_days()[1])),
#         'last_month': (format_date(get_last_month()[0]), format_date(get_last_month()[1])),
#         'table_name_filter': 'startupindia',
#     }

#     return render(request, 'fema/gridfilter.html', context)







    






def get_status_from_config(source_name):
    config_path = get_config_path(source_name)
    config = configparser.ConfigParser()
    config.read(config_path)
    return config.get(source_name, 'status')

def get_config_path(source_name):
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    config_folder = os.path.join(base_dir, 'config')
    config_file = f'config_{source_name.lower()}.ini'
    return os.path.join(config_folder, config_file)


# Function to get a unique Sr_no value
def get_unique_sr_no():
    max_sr_no = rbi_log.objects.using('rbi').aggregate(models.Max('Sr_no'))['Sr_no__max']
    return max_sr_no + 1 if max_sr_no is not None else 1



# def update_status_in_database():
#     fema_status = get_status_from_config('rbi_fema')
#     ecb_status = get_status_from_config('rbi_ecb')
#     odi_status = get_status_from_config('rbi_odi')
#     # startupindia_status = get_status_from_config('startupindia')

    
#     try:
#         # Get the latest entry for rbi_fema
#         latest_rbi_fema_entry = rbi_log.objects.using('rbi').filter(source_name='rbi_fema').latest('Sr_no')

#         # Update or create for rbi_fema
#         rbi_fema_entry, created = rbi_log.objects.using('rbi').filter(source_name='rbi_fema', Sr_no=latest_rbi_fema_entry.Sr_no).update_or_create(
#             source_name='rbi_fema',
#             Sr_no=latest_rbi_fema_entry.Sr_no,
#             defaults={'source_status': fema_status}
#         )

#         # If entry is not created and there is a change in status, update the Sr_no
#         if not created and rbi_fema_entry.source_status != fema_status:
#             rbi_fema_entry.source_status = fema_status
#             rbi_fema_entry.save()

#     except ObjectDoesNotExist:
#         # Handle the case where there is no entry with source_name='rbi_fema'
#         # For example, you might want to create a new entry in this case
#             rbi_fema_entry = rbi_log.objects.using('rbi').create(source_name='rbi_fema', Sr_no=get_unique_sr_no(), source_status=fema_status)
            
        
#     # Update or create for rbi_ecb
#     try:
#         latest_rbi_ecb_entry = rbi_log.objects.using('rbi').filter(source_name='rbi_ecb').latest('Sr_no')
#         rbi_ecb_entry, created_ecb = rbi_log.objects.using('rbi').filter(source_name='rbi_ecb', Sr_no=latest_rbi_ecb_entry.Sr_no).update_or_create(
#             source_name='rbi_ecb',
#             Sr_no=latest_rbi_ecb_entry.Sr_no,
#             defaults={'source_status': ecb_status}
#         )
#         if not created_ecb and rbi_ecb_entry.source_status != ecb_status:
#             rbi_ecb_entry.source_status = ecb_status
#             rbi_ecb_entry.save()

#     except ObjectDoesNotExist:
#         rbi_ecb_entry = rbi_log.objects.using('rbi').create(source_name='rbi_ecb', Sr_no=get_unique_sr_no(), source_status=ecb_status)


#     # Update or create for rbi_odi
#     try:
#         latest_rbi_odi_entry = rbi_log.objects.using('rbi').filter(source_name='rbi_odi').latest('Sr_no')
#         rbi_odi_entry, created_odi = rbi_log.objects.using('rbi').filter(source_name='rbi_odi', Sr_no=latest_rbi_odi_entry.Sr_no).update_or_create(
#             source_name='rbi_odi',
#             Sr_no=latest_rbi_odi_entry.Sr_no,
#             defaults={'source_status': odi_status}
#         )
#         if not created_odi and rbi_odi_entry.source_status != odi_status:
#             rbi_odi_entry.source_status = odi_status
#             rbi_odi_entry.save()

#     except ObjectDoesNotExist:
#         rbi_odi_entry = rbi_log.objects.using('rbi').create(source_name='rbi_odi', Sr_no=get_unique_sr_no(), source_status=odi_status)


#     # Update or create for startupindia
#     # try:
#     #     latest_startupindia_entry = startup_india_log.objects.using('startup_india').filter(source_name='startupindia').latest('Sr_no')
#     #     startupindia_entry, created_startupindia = startup_india_log.objects.using('startup_india').filter(source_name='startupindia', Sr_no=latest_startupindia_entry.Sr_no).update_or_create(
#     #         source_name='startupindia',
#     #         Sr_no=latest_startupindia_entry.Sr_no,
#     #         defaults={'source_status': startupindia_status}
#     #     )
#     #     if not created_startupindia and startupindia_entry.source_status != startupindia_status:
#     #         startupindia_entry.source_status = startupindia_status
#     #         startupindia_entry.save()

#     # except ObjectDoesNotExist:
#     #     startupindia_entry = startup_india_log.objects.using('rbi').create(source_name='startupindia', Sr_no=get_unique_sr_no(), source_status=startupindia_status)
        
