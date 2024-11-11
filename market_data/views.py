from .models import scraping_log
import pandas as pd
from django.http import HttpResponse, JsonResponse,request
from django.shortcuts import render
from datetime import datetime, date ,timedelta
from django.utils import timezone 
import json
from django.shortcuts import get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET,require_http_methods
from django.db.models import Case, When, Value, CharField
from openpyxl import Workbook
from openpyxl.styles import Font
from django.core.exceptions import ObjectDoesNotExist
from django.http import Http404
from django.http import HttpResponseForbidden


import logging



# views.py
from django.conf import settings
from django.http import HttpResponse

def debug_static(request):
    # static_root = settings.STATIC_ROOT
    static_dirs = settings.STATICFILES_DIRS
    return HttpResponse(f"""
    
        STATICFILES_DIRS: {static_dirs}
        DEBUG: {settings.DEBUG}
    """)


# Configure logging
logger = logging.getLogger(__name__)


def handler404(request, exception):
    """
    Custom 404 handler that logs the attempt and returns a user-friendly error page
    """
    # Log the 404 error with minimal information
    logger.warning(f"404 error for path: {request.path}")
    
    return render(request, 'errors/404.html', status=404)


# def handler403(request, exception):
#     return HttpResponseForbidden(render(request,'errors/403.html', status=403))
    
    # return render(request, 'errors/404.html', status=404)


# @require_http_methods(["GET"])
def dashboard(request):
    table_names = scraping_log.objects.values_list('table_name', flat=True).distinct()
    return render(request, 'dashboard.html', {'table_names': table_names})


# URL to Table name mapping
URL_TO_TABLE_MAPPING = {
    'company-master': 'ace_Company_master',
    'bse-equity': 'bse_new_equity',
    'company-equity': 'ace_company_equity',
    'company-equity-cons': 'ace_company_equity_cons',
    'ace-high-low': 'ace_52whl',
    'market-capital': 'bse_market_capital',
    'financial-prov': 'bse_financial_PROV',
    'financial-qc': 'bse_financial_QC',
    'finance-quarterly': 'ace_finance_Quarterly',
    'finance-quarterly-cons': 'ace_finance_Quarterly_Cons',
    'finance-cf': 'ace_Finance_cf',
    'finance-cons-cf': 'ace_Finance_cons_cf',
    'finance-fr': 'ace_Finance_fr',
    'finance-bs': 'ace_Finance_bs',
    'finance-cons-bs': 'ace_Finance_cons_bs',
    'capitaline-standalone': 'Capitaline_standalone',
    'capitaline-consolidated': 'Capitaline_consolidated',
    'bse-pledge': 'bse_pledge',
    'nse-pledge': 'nse_pledge_new',
    'bse-shareholding': 'bse_shp',
    'ace-shareholding': 'ace_shp',
    'nse-equity-list': 'NSE security list equity segment',
    'nse-sme-list': 'NSE security list trading sme',
    'bse-security-list': 'bse security list'
}

# Reverse mapping for converting table names to URLs
TABLE_TO_URL_MAPPING = {v: k for k, v in URL_TO_TABLE_MAPPING.items()}


# @require_http_methods(["GET"])
def table_details(request, url_name):

     # Convert URL name to actual table name
    table_name = URL_TO_TABLE_MAPPING.get(url_name)

    # if not table_name:
    #     logger.warning(f"Invalid URL access attempt: {url_name}")
    #     raise Http404(f"Invalid URL '{url_name}'")


    amber_table_names = [
        'ace_52whl',
        'ace_Company_master',
        'ace_company_master',
        'ace_Finance_fr',
        'ace_Finance_cf',
        'ace_company_equity',
        'ace_Finance_cons_cf',
        'ace_company_equity_cons',
        'ace_finance_Quarterly',
        'ace_finance_Quarterly_Cons',
        'ace_shp',
        'ace_Finance_bs',
        'ace_Finance_cons_bs',
    ]

    table_list = [
                  'ace_Company_master',
                  'bse_new_equity',
                  'ace_company_equity',
                  'ace_company_equity_cons',
                  'ace_52whl',
                  'bse_market_capital',
                  'bse_financial_PROV',
                  'bse_financial_QC',
                  'ace_finance_Quarterly',
                  'ace_finance_Quarterly_Cons',
                  'ace_Finance_cf',
                  'ace_Finance_cons_cf',
                  'ace_Finance_fr',
                  'ace_Finance_bs',
                  'ace_Finance_cons_bs',
                  'Capitaline_standalone',
                  'Capitaline_consolidated',
                  'bse_pledge',
                  'nse_pledge_new',
                  'bse_shp',
                  'ace_shp',
                  'NSE security list equity segment',
                  'NSE security list trading sme',
                  'bse security list'
                  ]
    
    # Check if the table_name is in the allowed list
    if table_name not in table_list:
        logger.warning(f"Invalid table access attempt: {table_name}")
        raise Http404(f"Table '{table_name}' does not exist")

    # print("debug======",table_names)
    # Fetch all unique table names
    table_names = scraping_log.objects.values_list('table_name', flat=True).distinct()


     # Create ordered url_friendly_names following table_list order
    url_friendly_names = []
    for table in table_list:
        if table in table_names:
            url_name = TABLE_TO_URL_MAPPING.get(table, table)
            url_friendly_names.append((table, url_name))
    print("\n\nurl_friendly_names==== ",  url_friendly_names)

    table_index = {name: index for index, name in enumerate(table_list)}

    # Use Case, When, and Value to order the queryset based on the custom order
    table_names = table_names.annotate(
        custom_order=Case(
            *[When(table_name=name, then=Value(index)) for name, index in table_index.items()],
            default=Value(len(table_index)),
            output_field=CharField()
        )
    ).order_by('custom_order')

    print("\n\ntable_names", table_names)
    # Check if the requested table_name is valid
    if table_name not in table_names:
        print("table names for debug ======",table_name)
        return render(request, 'table_details.html', {'error_message': f'Table {table_name} not found.'})

    # Fetch all data for the selected table
    table_data = scraping_log.objects.filter(table_name=table_name)
    

    # Order the data by trade_date in descending order
    table_data = table_data.order_by('-trade_date')

    # Handle date range selection
    time_range = request.GET.get('time_range')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    today = datetime.today().date()
    yesterday = today - timedelta(days=1)

    if time_range == '7':
        start_date = today - timedelta(days=7)
        end_date = yesterday  
    elif time_range == '15':
        start_date = today - timedelta(days=15)
        end_date = yesterday
    elif time_range == '30':
        start_date = today - timedelta(days=30)
        end_date = yesterday 
    elif from_date and to_date:
        start_date = datetime.strptime(from_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(to_date, '%Y-%m-%d').date()
    else:
        # Default to last 7 days if no specific range is selected
        start_date = today - timedelta(days=7)
        end_date = yesterday 
    # Filter data based on the date range
    if start_date is not None:
        table_data = table_data.filter(trade_date__range=[start_date,end_date])

    # Get unique failure reasons
    failure_reasons = table_data.filter(status='failure').values_list('reason', flat=True).distinct()

    # Handle date filter
    selected_date = request.GET.get('scraping_date') 
    if selected_date:
        # Convert the selected_date to the desired format
        selected_date = datetime.strptime(selected_date, '%b. %d, %Y, %I:%M %p').strftime('%Y-%m-%d')
        table_data = table_data.filter(trade_date=selected_date)

    # Create a list of all dates in the desired range
    date_range = [start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1)][::-1]

    # Create a list of dictionaries for each date, including '-' for dates without data
    structured_data = []

    # for date_str in date_range:
    #     data_entry = {
    #         'trade_date': date_str,
    #         'data': table_data.filter(trade_date=date_str).first(),
    #         'status': '-',
    #     }

    #     if data_entry['data']:
    #         data_entry['status'] = data_entry['data'].status.strip()

    #     structured_data.append(data_entry)

    # recent_status = get_recent_status(table_name = table_name)
    
    
    
    for date_str in date_range:
        # Filter data for the specific trade date and order by descending trade_date to get the most recent entry
        data_entries = table_data.filter(trade_date=date_str).order_by('-Scraped_on')

    # Get the most recent entry for the trade date
        most_recent_entry = data_entries.first()

        data_entry = {
            'trade_date': date_str,
            'data': most_recent_entry,
            'status': '-' if most_recent_entry is None else most_recent_entry.status.strip(),
        }


        if data_entry['data']:
            data_entry['status'] = data_entry['data'].status.strip()

        structured_data.append(data_entry)

    recent_status = get_recent_status(table_name = table_name)

    data = {
        'table_name': table_name,
        'url_name': url_name,
        'url_friendly_names': url_friendly_names,
        'structured_data': structured_data,
        'failure_reasons': failure_reasons,
        'table_names': table_names,
        'time_range': time_range,
        'from_date': from_date,
        'to_date': to_date,
        'scraping_date': selected_date,
        'amber_table_names': amber_table_names,
        'today': today,
        'end_date':end_date,
        'start_date':start_date,
        'yesterday': yesterday,
        'table_list':table_list,
        'recent_status' : recent_status,
    }
    if 'download_excel' in request.GET:
        time_range = request.GET.get('time_range')
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        
# Use the custom date range if specified, otherwise use the default logic
        if time_range == '7':
            start_date = today - timedelta(days=7)
            end_date = yesterday
        elif time_range == '15':
            start_date = today - timedelta(days=15)
            end_date = yesterday
        elif time_range == '30':
            start_date = today - timedelta(days=30)
            end_date = yesterday
        elif time_range == 'custom' and from_date and to_date:
            start_date = datetime.strptime(from_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(to_date, '%Y-%m-%d').date()
            # Default to last 7 days if no specific range is selected
        else:
            start_date = datetime.now().date() - timedelta(days=7)
            end_date = yesterday

        # Fetch all dates within the specified range
        all_dates = [start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1)]

        # Fetch data for the specified table and all dates within the range
        excel_data = scraping_log.objects.filter(table_name=table_name, trade_date__in=all_dates).order_by('-trade_date')
       # Create a dictionary to store data for each date
        date_data_dict = {entry.trade_date: entry for entry in excel_data}

# Add placeholder entries for dates without data
        for date in all_dates:
            if date not in date_data_dict:
            # Create a placeholder entry with default values
                placeholder_entry = scraping_log(trade_date=date, status='No Data', no_of_data_available=0, no_of_data_scraped=0, reason='No Data',newly_added_count=0,deleted_source_count=0)
                date_data_dict[date] = placeholder_entry

                filename = f'{table_name}_data_{start_date}_{end_date}.xlsx'
        # Create a response object with the appropriate content type for Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        # Construct the filename with the table_name and time_range_str
        response['Content-Disposition'] = f'attachment; filename={filename}'
        # Create a workbook and add a worksheet
        workbook = Workbook()
        worksheet = workbook.active

        # Set font style for headers
        header_font = Font(bold=True)

        # Add headers to the worksheet
        headers = ['Table Name', 'Status','#Records Available', '#Records Scraped','Failure Reason', 'Trade date', 'Newly Added Count','Deleted Source Count', 'Scraped on']
        for col_num, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font

        # Populate data into the worksheet
        for row_num, data_entry in enumerate(excel_data, start=2):
            worksheet.cell(row=row_num, column=1, value=data_entry.table_name)
            worksheet.cell(row=row_num, column=2, value=data_entry.status)
            worksheet.cell(row=row_num, column=3, value=data_entry.no_of_data_available)
            worksheet.cell(row=row_num, column=4, value=data_entry.no_of_data_scraped)
            worksheet.cell(row=row_num, column=5, value=data_entry.reason)
            worksheet.cell(row=row_num, column=6, value=data_entry.trade_date)
            worksheet.cell(row=row_num, column=7, value=data_entry.newly_added_count)
            worksheet.cell(row=row_num, column=8, value=data_entry.deleted_source_count)
            worksheet.cell(row=row_num, column=9, value=data_entry.Scraped_on)

        # Save the workbook to the response
        workbook.save(response)
        return response
    print("\n\n\n debugging the data========",data) 
    return render(request, 'table_details.html', data)

# @require_http_methods(["GET"])
def table_details2(request):
    time_range = request.GET.get('time_range')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    today = datetime.today().date()

    if time_range == '7':
        start_date = today - timedelta(days=7)
    else:
        start_date = today - timedelta(days=7)

    yesterday = today - timedelta(days=1)

    table_list = [
        'ace_Company_master',
        'bse_new_equity',
        'ace_company_equity',
        'ace_company_equity_cons',
        'ace_52whl',
        'bse_market_capital',
        'bse_financial_PROV',
        'bse_financial_QC',
        'ace_finance_Quarterly',
        'ace_finance_Quarterly_Cons',
        'ace_Finance_cf',
        'ace_Finance_cons_cf',
        'ace_Finance_fr',
        'ace_Finance_bs',
        'ace_Finance_cons_bs',
        'Capitaline_standalone',
        'Capitaline_consolidated',
        'bse_pledge',
        'nse_pledge_new',
        'bse_shp',
        'ace_shp',
        'NSE security list equity segment',
        'NSE security list trading sme',
        'bse security list'
        
    ]

    # Get unique table names for the sidebar and order them based on the table_list
    table_names = scraping_log.objects.values_list('table_name', flat=True).distinct()

    # Create a dictionary to map table names to their index in the custom order
    table_index = {name: index for index, name in enumerate(table_list)}

    # Use Case, When, and Value to order the queryset based on the custom order
    table_names = table_names.annotate(
        custom_order=Case(
            *[When(table_name=name, then=Value(index)) for name, index in table_index.items()],
            default=Value(len(table_index)),
            output_field=CharField()
        )
    ).order_by('custom_order')

    amber_table_names = [
        'ace_52whl',
        'ace_Company_master',
        'ace_company_master',
        'ace_Finance_fr',
        'ace_Finance_cf',
        'ace_company_equity',
        'ace_Finance_cons_cf',
        'ace_company_equity_cons',
        'ace_finance_Quarterly',
        'ace_finance_Quarterly_Cons',
        'ace_shp',
        'ace_Finance_bs',
        'ace_Finance_cons_bs',
    ]

    failure_reasons = scraping_log.objects.filter(status='failure').values_list('reason', flat=True).distinct()


    # Initialize structured_data dictionary
    structured_data = {}

    recent_statuses = {}

    # Iterate through each source name
    for table_name in table_list:
        # Get the recent status for the source
        recent_status = get_recent_status(table_name=table_name)
        recent_statuses[table_name] = recent_status

    if start_date is not None:
        date_range = [start_date + timedelta(days=i) for i in range((today - start_date).days)][::-1]
    else:
        date_range = []

    # Iterate through each table name
    for table_name in table_names:
        # Filter records for the specific table and date range
        if start_date is not None:
            records = scraping_log.objects.filter(table_name=table_name).order_by('-Scraped_on')
            # print("records debug========".records)
        else:
            records = scraping_log.objects.filter(table_name=table_name).order_by('-Scraped_on')
            # print("records debug2222========".records)

        # Create a dictionary to store data for each date
        date_data = {}

        # Populate the dictionary with actual data
        for date in date_range:
            record = records.filter(trade_date__exact=str(date)).first()
            
            if record:
                # print("for debug---------",record)
                record.status_stripped = record.status.strip()
                date_data[date] = {
                    'no_of_data_available': record.no_of_data_available,
                    'no_of_data_scraped': record.no_of_data_scraped,
                    'status': record.status_stripped,
                    'Scraped_on': record.Scraped_on,
                    'Table Name': record.table_name,
                    'Reason': record.reason,
                    'Trade Date': record.trade_date,
                    'total_record_count': record.total_record_count,
                    'source_status':record.source_status,
                    'Newly Added Count': record.newly_added_count,
                    'Deleted Source Count': record.deleted_source_count
                }
                
            else:
                date_data[date] = {
                    'no_of_data_available': '-',
                    'no_of_data_scraped': '-',
                    'Scraped_on': '-',
                    'status': '-',
                    'Table Name': '-',
                    'Reason': '-',
                    'Trade Date': '-',
                    'total_record_count': 0,
                    'source_status': 'Active',
                    'newly_added_count' : '-',
                    'deleted_source_count': '-'
                }

        # Append table data to the main structured_data dictionary
        structured_data[table_name] = date_data
    
    # Create URL-friendly names for template
    url_friendly_names = [(name, TABLE_TO_URL_MAPPING.get(name, name)) for name in table_names]
    print("url_friendly_names :::::: ",  url_friendly_names)

    # Context data for rendering the HTML template
    context = {
        'table_names': table_names,
        'date_range': date_range,
        'structured_data': structured_data,
        'amber_table_names': amber_table_names,
        'failure_reasons': failure_reasons,
        'today': today,
        'start_date': start_date,
        'yesterday': yesterday,
        'table_list': table_list,
        'source_status': recent_statuses,
        'url_friendly_names': url_friendly_names, 
    }
    return render(request, 'pivottable.html', context)


# @require_http_methods(["GET"])
def get_data_for_popup(request, table_name):
    today = datetime.today().date()
    yesterday = today - timedelta(days=1)

    data = scraping_log.objects.filter(table_name=table_name,trade_date=yesterday).order_by('-trade_date', '-Scraped_on').first()
    print("output=======",data)

    amber_table_names = [
        'ace_52whl',
        'ace_company_master',
        'ace_Finance_fr',
        'ace_Finance_cf',
        'ace_company_equity',
        'ace_Finance_cons_cf',
        'ace_company_equity_cons',
        'ace_finance_Quarterly',
        'ace_finance_Quarterly_Cons',
        'ace_shp',
        'ace_Finance_bs',
        'ace_Finance_cons_bs',
    ]

    if data:
        # Format the timestamp into a string with a specific format
        formatted_scraped_on = data.Scraped_on.strftime("%d-%m-%Y")
        # Assuming data.trade_date is a string in the format "2023-12-29"
        trade_date_str = data.trade_date

# Convert the string to a datetime object
        trade_date = datetime.strptime(trade_date_str, "%Y-%m-%d")

# Format the datetime object as "29-12-2023"
        formatted_trade_date = trade_date.strftime("%d-%m-%Y")

        response_data = {
            'table_name': data.table_name,
            'status': data.status,
            'no_of_data_scraped': data.no_of_data_scraped,
            'reason': data.reason,
            'amber_table_names': amber_table_names,
            'trade_date': formatted_trade_date,
            'Scraped_on': formatted_scraped_on,  # Use the formatted timestamp
        }
        return JsonResponse(response_data)
    else:
        return JsonResponse({'message': 'Data not available for today.'})
    
    
# @require_http_methods(["GET"])
def get_recent_status(table_name):
    try:
        # Fetch the recent status for today's date
        recent_status_entry = scraping_log.objects.filter(table_name=table_name).latest('trade_date')
        recent_status = recent_status_entry.source_status if recent_status_entry is not None and recent_status_entry.source_status else None
        if recent_status is not None:
            return recent_status

        # If today's status is null, iterate over previous dates until a non-null status is found
        for i in range(1,100):  # Assuming checking for the past 99 days
            previous_date = timezone.now().date() - timedelta(days=i)
            previous_status_entry = scraping_log.objects.filter(table_name=table_name, trade_date = previous_date).first()
            if previous_status_entry is not None and previous_status_entry.source_status:
                return previous_status_entry.source_status

    except ObjectDoesNotExist:
        pass

    return 'N/A'  # Return 'N/A' if no status is found
